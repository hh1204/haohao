import openpyxl

# 打开excel表格
workbook = openpyxl.load_workbook(r'D:\python3\test_unittest\data\hao.xlsx')

# 获取所有的表单
sheet = workbook.sheetnames
print(sheet)
"""获取表单的4种方式"""
# 第一种，通过索引获取
sheet1 = workbook.worksheets[0]
print(sheet1)
# 第二种，通过get_sheet_by_name()方法获取,这种会出现警告。
sheet2 = workbook.get_sheet_by_name('Sheet1')
print(sheet2)
# 第三种，通过表单名直接获取，类似于字典
sheet3 = workbook['Sheet1']
print(sheet3)
# 第四种，获取激活的表单，也就是正在使用的表单，这种方式不推荐使用，了解一下就好
sheet4 = workbook.active
print(sheet4)
# 获取单元格的值，注意：单元格的值是以1开头，不是0
cell1 = sheet1.cell(1,1).value
print(cell1)

# 获取一行的数据
a = sheet1[1]
print(a)
# 获取多行的数据，注意：要和切片区分开，这里左右都是能取到的，切片是取左不取右
a1 = sheet1[1:3]
print(a1)
# 获取一列的数据
b = sheet1["B"]
print(b)

data_row = list(sheet1.rows)[0:]
data = []
for row in data_row:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    data.append(row_data)

print(data)
datas = []
max_row = sheet1.max_row
max_column = sheet1.max_column
for row in range(2,max_row+1):
    row_data = []
    for column in range(1,max_column):
        row_data.append(sheet1.cell(row,column).value)
    datas.append(row_data)
print(datas)

# 写,写的单元格有值就是修改
sheet1.cell(1,1).value = 'girl'
sheet1.cell(1,3).value = '女孩'

# 保存
# workbook.save(r'D:\python3\test_unittest\data\hao.xlsx')
# 关闭
workbook.close()