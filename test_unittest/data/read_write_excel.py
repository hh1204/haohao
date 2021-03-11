
from openpyxl import load_workbook


class Excel:
    """excel 封装"""
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = load_workbook(file_name)

    def open_excel(self):
        """打开excel和获取表单"""
        self.workbook = load_workbook(self.file_name)
        if isinstance(self.sheet_name, int):
            self.sheet = self.workbook.worksheets[self.sheet_name]
        else:
            self.sheet = self.workbook[self.sheet_name]

    # def choose_sheet(self, sheet_name):
    #     """选择表单.
    #     sheet_name 是整数，根据索引获取。
    #     如果是字符串，根据名字获取 'haohao'
    #     """
    #     if isinstance(sheet_name, int):
    #         return self.wb.worksheets[sheet_name]
    #     return self.wb[sheet_name]

    def read_headers(self):
        """获取标题"""
        self.open_excel()
        headers = [c.value for c in self.sheet[1]]
        self.wb.close()
        return headers

    def read_data(self,start_row=2, start_column=1):
        """获取所有的数据"""
        self.open_excel()
        sheet = self.sheet
        header = [c.value for c in sheet[1]]
        data = []
        for row in range(start_row, sheet.max_row+1):
            row_data = []
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            row_data = dict(zip(header, row_data))
            data.append(row_data)
        self.workbook.close()
        return data

    def save(self):
        """保存和关闭"""
        self.workbook.save(self.file_name)
        self.workbook.close()

    def write(self, row, column, data):
        """
        写数据或者修改
        :param row:
        :param column:
        :param data:
        :return:
        """
        self.open_excel()
        self.sheet.cell(row, column).value = data
        self.save()


class Excel2:
    """excel 封装"""
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = load_workbook(file_name)
        # if isinstance(sheet_name, int):
        #     self.sheet =  self.wb.worksheets[sheet_name]
        # else:
        #     self.sheet = self.wb.get_sheet_by_name(sheet_name)

    def choose_sheet(self, sheet_name):
        """选择表单.
        sheet_name 是整数，根据索引获取。
        如果是字符串，根据名字获取 'haohao'
        """
        if isinstance(sheet_name, int):
            return self.workbook.worksheets[sheet_name]
        return self.workbook[sheet_name]

    def read_line(self, sheet_name, line):
        """获取行"""
        sheet = self.choose_sheet(sheet_name)
        sheet_data =  sheet[line]
        # 元组 （Cell(1,1), Cell(1,2）
        data = []
        for c in sheet_data:
            data.append(c.value)
        return data

    def read_data(self, sheet_name, start_row=2, start_column=1):
        """获取所有的数据"""
        sheet = self.choose_sheet(sheet_name)
        # max_row, max column
        data = []
        for row in range(start_row, sheet.max_row + 1):
            row_data = []
            for column in range(start_column, sheet.max_column + 1):
                row_data.append(sheet.cell(row, column).value)
            data.append(row_data)
        return data

    def read_cell(self, sheet_name, row, column):
        """一个单元格的数据"""
        sheet = self.choose_sheet(sheet_name)
        return sheet.cell(row, column).value

    def save(self):
        """保存和关闭"""
        self.workbook.save(self.file_name)
        self.workbook.close()

    @staticmethod
    def write(file_name, sheet_name, row, column, data):
        workbook = load_workbook(file_name)
        sheet = workbook.get_sheet_by_name(sheet_name)
        sheet.cell(row, column).value = data
        # 保存关闭
        workbook.save(file_name)
        workbook.close()


if __name__ == '__main__':
    test = Excel(r'D:\python3\test_unittest\data\hao.xlsx','Sheet1')
    print(test.read_data())